from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


OUTPUT = Path("/Users/youyang/项目/2026/商品中台/图书多渠道商品中台-功能清单（Excel版）.xlsx")
TEMPLATE_OUTPUT = Path("/Users/youyang/项目/2026/商品中台/图书多渠道商品中台-导入模板字段（Excel版）.xlsx")


FUNCTION_ROWS = [
    ("产品管理", "产品列表", "必做", "列表页", "菜单直达", "查看和筛选产品主档信息，作为一期产品维护主入口。"),
    ("产品管理", "产品详情", "必做", "详情页", "列表内容进入", "维护产品基础信息、默认内容、素材关系和 SKU 摘要信息。"),
    ("产品管理", "产品导入", "必做", "工作页", "二级菜单直达", "基于导入模板导入产品与商品资料，完成校验、关联和确认导入。"),
    ("产品管理", "导入记录", "必做", "记录页", "导入页进入", "查看导入批次、成功失败结果和异常原因，支持回溯导入处理结果。"),
    ("产品管理", "商家编码关联", "必做", "规则能力", "导入处理过程", "通过商家编码建立产品与多渠道多店铺商品的关联关系，支撑同一产品下不同商品售卖信息统一归集。"),
    ("产品管理", "新增产品", "增强", "功能动作", "列表按钮进入", "支持少量手工新增产品主档，补充模板导入之外的产品资料。"),
    ("素材管理", "素材列表", "增强", "列表页", "菜单直达", "统一查看导入后的主图、详情图和 SKU 图素材，支持预览和补充维护。"),
    ("素材管理", "素材详情", "增强", "详情页", "列表内容进入", "查看素材预览、路径和引用关系，支持替换和重新绑定。"),
    ("商品管理", "渠道商品列表", "必做", "列表页", "菜单直达", "统一查看同一产品在多渠道多店铺下的不同商品售卖信息，支持按渠道、店铺、状态筛选。"),
    ("商品管理", "渠道商品详情", "必做", "详情页", "列表内容进入", "轻量维护渠道商品标题、价格、素材、类目属性、销售属性、上下架状态等差异内容。"),
    ("商品管理", "店铺级类目动态适配", "必做", "规则能力", "详情页联动", "商品类目需根据所选渠道和店铺动态调整，并联动类目、类目属性、销售属性及可维护字段范围。"),
    ("商品管理", "批量修改", "增强", "功能动作", "二级菜单直达", "提供批量修改能力，支持批量修改标题、价格、批量上下架和库存。"),
    ("库存管理", "库存列表", "必做", "列表页", "菜单直达", "统一查看并维护同一产品在各渠道商品和各店铺下的库存总览。"),
    ("库存管理", "库存详情", "增强", "详情页", "列表内容进入", "查看单条库存记录的总库存、可用库存、锁定库存和可售库存，支持调整与同步。"),
    ("库存管理", "库存调整记录", "增强", "记录页", "二级菜单直达", "查看库存台账、调整前后变化、原因、操作人和时间等历史记录。"),
    ("配置管理", "配置列表", "必做", "列表页", "菜单直达", "汇总店铺级规则配置入口，作为一期配置主入口。"),
    ("配置管理", "店铺配置详情", "必做", "详情页", "列表内容进入", "维护店铺级规则，包括类目动态适配、运费模板、上下架状态和发货时效等配置。"),
]


PAGE_TREE_ROWS = [
    ("产品管理", "菜单入口", "产品列表", "列表页", "左侧菜单直达"),
    ("产品管理", "产品列表", "产品详情", "详情页", "列表内容进入"),
    ("产品管理", "菜单入口", "产品导入", "工作页", "二级菜单直达"),
    ("产品管理", "产品导入", "导入记录", "记录页", "导入页进入"),
    ("产品管理", "产品导入", "商家编码关联", "规则能力", "导入校验和落库过程中执行"),
    ("产品管理", "产品列表", "新增产品", "功能动作", "列表按钮进入"),
    ("素材管理", "菜单入口", "素材列表", "列表页", "左侧菜单直达"),
    ("素材管理", "素材列表", "素材详情", "详情页", "列表内容进入"),
    ("商品管理", "菜单入口", "渠道商品列表", "列表页", "左侧菜单直达"),
    ("商品管理", "渠道商品列表", "渠道商品详情", "详情页", "列表内容进入"),
    ("商品管理", "渠道商品详情", "店铺级类目动态适配", "规则能力", "按所选渠道和店铺动态联动类目与属性"),
    ("商品管理", "菜单入口", "批量修改", "功能动作", "二级菜单直达"),
    ("库存管理", "菜单入口", "库存列表", "列表页", "左侧菜单直达"),
    ("库存管理", "库存列表", "库存详情", "详情页", "列表内容进入"),
    ("库存管理", "菜单入口", "库存调整记录", "记录页", "二级菜单直达"),
    ("配置管理", "菜单入口", "配置列表", "列表页", "左侧菜单直达"),
    ("配置管理", "配置列表", "店铺配置详情", "详情页", "列表内容进入"),
]


IMPORT_TEMPLATE_FIELD_ROWS = [
    ("基础维度", "渠道", "必填", "用于标识目标发品渠道。"),
    ("基础维度", "店铺", "必填", "用于标识目标发品店铺，并联动店铺级类目规则。"),
    ("基础维度", "类目", "必填", "按渠道和店铺动态返回的可用发品类目。"),
    ("基础维度", "商品ID", "选填", "渠道侧商品标识，用于更新已有商品或建立映射。"),
    ("关联标识", "商家编码", "必填", "用于建立产品与商品的关联关系，是统一归集的核心关联键。"),
    ("类目属性", "类目属性", "必填", "按店铺级类目规则动态返回，作为发品必填属性集合。"),
    ("类目属性", "销售属性", "必填", "用于定义 SKU 维度的销售规格和属性值。"),
    ("履约信息", "运费模板", "必填", "按店铺规则可选，用于发品履约配置。"),
    ("履约信息", "上下架状态", "必填", "控制商品导入后的默认售卖状态。"),
    ("履约信息", "发货时效", "必填", "用于维护店铺级发货承诺和履约时效。"),
    ("SKU信息", "SKU信息", "必填", "包含 SKU 编码、名称、规格、商家编码等 SKU 级数据。"),
    ("SKU信息", "库存", "必填", "用于维护各商品 SKU 的初始库存或可售库存。"),
    ("图片素材", "主图", "必填", "商品主图信息，用于商品展示。"),
    ("图片素材", "详情", "必填", "商品详情图或详情描述信息。"),
    ("图片素材", "SKU图", "选填", "用于维护 SKU 级图片信息。"),
]


HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
HEADER_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="000000")
BODY_FONT = Font(name="Microsoft YaHei", size=11, color="000000")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def merge_same_values(ws, rows, tuple_index, column_index, start_row=2):
    current_value = None
    group_start = start_row

    for row_idx, row in enumerate(rows, start=start_row):
        value = row[tuple_index]
        if current_value is None:
            current_value = value
            group_start = row_idx
            continue
        if value != current_value:
            if row_idx - 1 > group_start:
                ws.merge_cells(start_row=group_start, start_column=column_index, end_row=row_idx - 1, end_column=column_index)
            current_value = value
            group_start = row_idx

    if rows:
        last_row = start_row + len(rows) - 1
        if last_row > group_start:
            ws.merge_cells(start_row=group_start, start_column=column_index, end_row=last_row, end_column=column_index)


def apply_sheet_style(ws, headers, max_row, left_align_columns):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGNMENT

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = BORDER
            cell.font = BODY_FONT
            cell.alignment = LEFT_ALIGNMENT if cell.column in left_align_columns else CENTER_ALIGNMENT

    for row_idx in range(1, max_row + 1):
        ws.row_dimensions[row_idx].height = 34 if row_idx == 1 else 30


def build_function_sheet(wb):
    ws = wb.active
    ws.title = "一期功能清单"
    headers = ["一级模块", "页面/功能名称", "一期分层", "类型", "入口方式", "功能说明"]

    for row_idx, row_values in enumerate(FUNCTION_ROWS, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    merge_same_values(ws, FUNCTION_ROWS, tuple_index=0, column_index=1)
    max_row = len(FUNCTION_ROWS) + 1
    apply_sheet_style(ws, headers, max_row, left_align_columns={6})

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 66


def build_page_tree_sheet(wb):
    ws = wb.create_sheet("页面树")
    headers = ["一级模块", "上级节点", "当前节点", "节点类型", "入口说明"]

    for row_idx, row_values in enumerate(PAGE_TREE_ROWS, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    merge_same_values(ws, PAGE_TREE_ROWS, tuple_index=0, column_index=1)
    merge_same_values(ws, PAGE_TREE_ROWS, tuple_index=1, column_index=2)
    max_row = len(PAGE_TREE_ROWS) + 1
    apply_sheet_style(ws, headers, max_row, left_align_columns={5})

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 32


def build_import_template_fields_sheet(wb):
    ws = wb.active
    ws.title = "导入模板字段"
    headers = ["字段分组", "模板字段", "是否必填", "字段说明"]

    for row_idx, row_values in enumerate(IMPORT_TEMPLATE_FIELD_ROWS, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    merge_same_values(ws, IMPORT_TEMPLATE_FIELD_ROWS, tuple_index=0, column_index=1)
    max_row = len(IMPORT_TEMPLATE_FIELD_ROWS) + 1
    apply_sheet_style(ws, headers, max_row, left_align_columns={4})

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 58


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_function_sheet(wb)
    build_page_tree_sheet(wb)
    wb.save(OUTPUT)

    template_wb = Workbook()
    build_import_template_fields_sheet(template_wb)
    template_wb.save(TEMPLATE_OUTPUT)


if __name__ == "__main__":
    main()
